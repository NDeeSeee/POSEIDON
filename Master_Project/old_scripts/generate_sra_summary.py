#!/usr/bin/env python3
"""
Generate summary table of SRA metadata files.
Checks all xlsx files and counts samples per category.
"""

import openpyxl
from pathlib import Path

# Define paths
SRA_METADATA_DIR = Path(__file__).parent.parent / "SRAMetadataFiles"
OUTPUT_FILE = SRA_METADATA_DIR / "summary_table.md"

# Target sheet names to check
TARGET_SHEETS = ["Tumors", "Controls", "Bulk_CellTypes", "Premalignant"]

def count_samples(worksheet):
    """Count samples in worksheet (excluding header row)."""
    max_row = worksheet.max_row
    # Subtract 1 for header row, but only if there's data
    return max(0, max_row - 1) if max_row > 0 else 0

def process_xlsx_file(filepath):
    """Process single xlsx file and return summary data."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        
        # Get cancer type from filename
        cancer_type = filepath.stem
        
        # Count samples for target sheets
        counts = {}
        for sheet_name in TARGET_SHEETS:
            if sheet_name in sheet_names:
                counts[sheet_name] = count_samples(wb[sheet_name])
            else:
                counts[sheet_name] = "None"
        
        # Find other sheets (not in target list)
        other_sheets = [s for s in sheet_names if s not in TARGET_SHEETS]
        
        wb.close()
        return cancer_type, counts, other_sheets
    
    except Exception as e:
        print(f"Error processing {filepath}: {e}")
        return None, None, None

def generate_summary():
    """Generate summary table for all xlsx files."""
    # Collect data
    data = []
    
    # Process all xlsx files
    xlsx_files = sorted(SRA_METADATA_DIR.glob("*.xlsx"))
    
    for xlsx_file in xlsx_files:
        cancer_type, counts, other_sheets = process_xlsx_file(xlsx_file)
        if cancer_type:
            data.append({
                'cancer_type': cancer_type,
                'counts': counts,
                'other_sheets': other_sheets
            })
    
    # Generate markdown table
    lines = []
    lines.append("# SRA Metadata Summary\n")
    lines.append(f"*Generated from {len(data)} xlsx files*\n")
    
    # Main table header
    lines.append("| Cancer Type | Tumors | Controls | Bulk_CellTypes | Premalignant | Other Sheets |")
    lines.append("|-------------|--------|----------|----------------|--------------|--------------|")
    
    # Data rows
    for item in data:
        cancer = item['cancer_type']
        counts = item['counts']
        others = ", ".join(item['other_sheets']) if item['other_sheets'] else "-"
        
        row = f"| {cancer} | {counts['Tumors']} | {counts['Controls']} | {counts['Bulk_CellTypes']} | {counts['Premalignant']} | {others} |"
        lines.append(row)
    
    # Write to file
    output_content = "\n".join(lines) + "\n"
    OUTPUT_FILE.write_text(output_content)
    
    print(f"? Summary generated: {OUTPUT_FILE}")
    print(f"  Processed {len(data)} files")

if __name__ == "__main__":
    generate_summary()
